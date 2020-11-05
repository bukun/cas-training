from openpyxl import load_workbook


def method_name():
    xlsx_obj = load_workbook('./demo_tmpl/e_mod.xlsx')
    show_info(xlsx_obj)

    work_sheet = xlsx_obj.active
    show_sheet_info(work_sheet)
    data_array = get_data(work_sheet)
    avg = sum(data_array) / len(data_array)
    print(avg)
    work_sheet.cell(row=1, column=10).value = avg
    xlsx_obj.save('xx_data.xlsx')


def get_data(work_sheet):
    if work_sheet is None:
        xlsx_obj = load_workbook('./demo_tmpl/e_mod.xlsx')
        work_sheet = xlsx_obj.active
    # for row in work_sheet.rows:
    #     # print(row)
    #     pass
    data_array = []
    for row_idx in range(4, work_sheet.max_row + 1):
        # print(sheet.cell(row = row_idx, column = 1).value)
        data_array.append(work_sheet.cell(row=row_idx, column=2).value)
    return data_array

def show_sheet_info(ws):
    # 查看工作表
    print('=' * 10)
    b4 = ws['B4']
    print(b4.value)
    b_foo = ws.cell(row = 4, column = 2)
    print(b_foo.value)

    print('=' * 10)

def show_info(xlsx_obj):
    # 查看工作薄

    work_sheet = xlsx_obj.get_sheet_by_name('Sheet1')
    print(work_sheet.max_row)
    print(work_sheet.max_column)
    return work_sheet


if __name__ == '__main__':
    method_name()