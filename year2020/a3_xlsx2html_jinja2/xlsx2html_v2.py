from jinja2 import FileSystemLoader, Environment

env = Environment(loader=FileSystemLoader('templates'))
template = env.get_template('html_base_v3.jinja2')

import openpyxl

wb = openpyxl.load_workbook('imgs/img_contents.xlsx')


sheet = wb.active


html_str = []

for i in range(2, sheet.max_row + 1):
    img_title = sheet.cell(row=i, column=1).value
    img_misc = sheet.cell(row=i, column=2).value
    img_length = sheet.cell(row=i, column=3).value
    img_width = sheet.cell(row=i, column=4).value

    if img_title:
        pass
    else:
        continue

    html_str.append([img_title, img_misc])


template.render(content=html_str)

with open('xx_html.html', 'w') as fo:
    fo.write(template.render(content=html_str))
